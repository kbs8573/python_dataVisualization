import openpyxl

# 데이터를 읽어올 파일과 Sheet를 각각 workbook, worksheet로 설정한다.
fileName = "C:\PythonExcercise\dataVisualization\월매출_Sample.xlsx".replace("\\", "/")
wb = openpyxl.load_workbook(fileName)
ws = wb["매출량"]

# 생성하고자 하는 차트를 정의한다.
from openpyxl.chart import LineChart, Reference
line_chart = LineChart()
line_chart.title = "월 매출량 추이"
line_chart.x_axis.title = "월"
line_chart.y_axis.title = "매출량"

# chart에 입력할 data를 정의하고, chart에 data를 입력한다.
data = Reference(ws, min_row=ws.min_row, max_row=ws.max_row, min_col=2, max_col=ws.max_column)
line_chart.add_data(data, titles_from_data=True)

categories = Reference(ws, min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
line_chart.set_categories(categories)

# worksheet에 chart를 작성하고, workbook을 저장해 준다.
ws.add_chart(line_chart, "A"+ str(ws.max_row + 2))
wb.save(fileName)