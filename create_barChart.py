import openpyxl
from openpyxl.chart import BarChart, Reference

fileName = "C:\PythonExcercise\dataVisualization\월매출_Sample.xlsx".replace("\\", "/")
wb = openpyxl.load_workbook(fileName)
ws = wb["매출량"]

# 생성하고자 하는 차트를 정의한다.
bar_chart = BarChart()
bar_chart.title = "월별 매출 및 순이익"
bar_chart.x_axis.title = "월"
bar_chart.y_axis.title = "금액"

# chart에 입력할 data를 정의하고, chart에 data를 입력한다.
data = Reference(ws, min_row=ws.min_row, max_row=ws.max_row, min_col=2, max_col=ws.max_column)
bar_chart.add_data(data, titles_from_data=True)

categories = Reference(ws, min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
bar_chart.set_categories(categories)

# worksheet에 chart를 작성하고, workbook을 저장해 준다.
ws.add_chart(bar_chart, "E2")
wb.save(fileName)